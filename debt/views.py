from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.db.models import Case, DecimalField, ExpressionWrapper, F, Q, Sum, When
from django.db.models.functions import Abs, Coalesce, Round
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DeleteView, ListView, TemplateView

from accounts.models import Customer
from .forms import EntryPaymentForm, SettlementForm
from .models import AccountType, DebtEntry, Settlement


# --- Helpers ---

def get_customer_debt_stats(customer_id, start_date=None):
    """
    Calculates detailed debt statistics for a customer.
    Returns a dictionary with raw values and the net balance.
    """
    entries = DebtEntry.objects.filter(customer_id=customer_id)
    if start_date:
        entries = entries.filter(entry_date__gte=start_date)

    # Receivable (Họ nợ mình)
    t_rec = entries.filter(account_type=AccountType.RECEIVABLE, is_settlement=False).aggregate(Sum('amount'))['amount__sum'] or 0
    p_rec = entries.filter(account_type=AccountType.RECEIVABLE, is_settlement=True).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Payable (Mình nợ họ)
    t_pay = entries.filter(account_type=AccountType.PAYABLE, is_settlement=False).aggregate(Sum('amount'))['amount__sum'] or 0
    p_pay = entries.filter(account_type=AccountType.PAYABLE, is_settlement=True).aggregate(Sum('amount'))['amount__sum'] or 0
    
    net = round((t_rec - p_rec) - (t_pay - p_pay), 0)
    
    return {
        'receivable_raw_total': t_rec,
        'receivable_paid': p_rec,
        'receivable_balance': t_rec - p_rec, # Độc lập phía Phải thu
        
        'payable_raw_total': t_pay,
        'payable_paid': p_pay,
        'payable_balance': t_pay - p_pay, # Độc lập phía Phải trả
        
        'net_balance': net # Chỉ cấn trừ ở dư nợ tổng
    }


def allocate_settlement_fifo(settlement, amount_to_allocate):
    """
    Logic for First-In-First-Out debt settlement.
    Allocates the paid amount to the oldest outstanding debt entries.
    """
    remaining = Decimal(str(amount_to_allocate))
    
    unpaid_entries = DebtEntry.objects.filter(
        customer=settlement.customer, 
        account_type=settlement.account_type, 
        is_settlement=False
    ).annotate(
        paid_so_far=Coalesce(Sum('payments__amount'), 0, output_field=DecimalField())
    ).annotate(
        rem=ExpressionWrapper(F('amount') - F('paid_so_far'), output_field=DecimalField())
    ).filter(rem__gt=0).order_by('entry_date', 'created_at')

    for entry in unpaid_entries:
        if remaining <= 0:
            break
        allocation = min(entry.rem, remaining)
        DebtEntry.objects.create(
            customer=settlement.customer, 
            account_type=settlement.account_type,
            parent_entry=entry, 
            settlement=settlement, 
            amount=allocation,
            is_settlement=True, 
            note=f"Khấu trừ nợ FIFO. Settlement #{settlement.id}",
            entry_date=settlement.payment_date
        )
        remaining -= allocation

    if remaining > 0:
        DebtEntry.objects.create(
            customer=settlement.customer, 
            account_type=settlement.account_type,
            settlement=settlement, 
            amount=remaining, 
            is_settlement=True,
            note=f"Tiền dư sau khi tất toán các đơn hàng. Settlement #{settlement.id}",
            entry_date=settlement.payment_date
        )
    return remaining


# --- Views ---

class DebtOverviewView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Main dashboard for viewing all customer balances."""
    template_name = 'debt/overview.html'
    context_object_name = 'customer_list'
    paginate_by = 50

    def test_func(self):
        return self.request.user.role == 'ADMIN'

    def handle_no_permission(self):
        messages.error(self.request, "Bạn không có quyền xem thông tin công nợ.")
        return redirect('dashboard')

    def get_queryset(self):
        queryset = Customer.objects.annotate(
            t_rec_raw=Coalesce(Sum('debt_entries__amount', filter=Q(debt_entries__account_type=AccountType.RECEIVABLE, debt_entries__is_settlement=False)), 0, output_field=DecimalField()),
            p_rec_raw=Coalesce(Sum('debt_entries__amount', filter=Q(debt_entries__account_type=AccountType.RECEIVABLE, debt_entries__is_settlement=True)), 0, output_field=DecimalField()),
            t_pay_raw=Coalesce(Sum('debt_entries__amount', filter=Q(debt_entries__account_type=AccountType.PAYABLE, debt_entries__is_settlement=False)), 0, output_field=DecimalField()),
            p_pay_raw=Coalesce(Sum('debt_entries__amount', filter=Q(debt_entries__account_type=AccountType.PAYABLE, debt_entries__is_settlement=True)), 0, output_field=DecimalField()),
        ).annotate(
            balance=Round(ExpressionWrapper(
                (F('t_rec_raw') - F('p_rec_raw')) - (F('t_pay_raw') - F('p_pay_raw')),
                output_field=DecimalField()
            ), 0)
        ).annotate(
            abs_balance=Abs('balance')
        ).order_by('-abs_balance')
        
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(name__icontains=q)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Global totals calculation
        annotated = self.get_queryset()
        totals = annotated.aggregate(
            total_rec=Sum(Case(When(balance__gt=0, then=F('balance')), default=0, output_field=DecimalField())),
            total_pay=Sum(Case(When(balance__lt=0, then=Abs(F('balance'))), default=0, output_field=DecimalField())),
        )
        context.update({
            'total_receivable': totals['total_rec'] or 0,
            'total_payable': totals['total_pay'] or 0,
            'search_query': self.request.GET.get('q', '')
        })
        return context


class CustomerDebtDetailView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Detailed ledger for a single customer with hierarchical grouping."""
    model = DebtEntry
    template_name = 'debt/customer_debt.html'
    context_object_name = 'entries'

    def test_func(self):
        return self.request.user.role == 'ADMIN'

    def get_queryset(self):
        customer_id = self.kwargs['customer_id']
        qs = DebtEntry.objects.filter(customer_id=customer_id)
        days = self.request.GET.get('days')
        if days and days.isdigit():
            start_date = date.today() - timedelta(days=int(days))
            qs = qs.filter(entry_date__gte=start_date)
        return qs.order_by('-entry_date', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer_id = self.kwargs['customer_id']
        customer = get_object_or_404(Customer, pk=customer_id)
        
        # 1. Grouping Logic for Timeline
        grouped_entries = []
        last_settlement_id = None
        for entry in self.get_queryset():
            if entry.settlement_id:
                if entry.settlement_id == last_settlement_id:
                    grouped_entries[-1]['sub_entries'].append(entry)
                else:
                    grouped_entries.append({
                        'type': 'settlement',
                        'settlement': entry.settlement,
                        'sub_entries': [entry],
                        'entry_date': entry.entry_date or entry.created_at
                    })
                    last_settlement_id = entry.settlement_id
            else:
                grouped_entries.append({
                    'type': 'solo',
                    'entry': entry,
                    'entry_date': entry.entry_date or entry.created_at
                })
                last_settlement_id = None
        
        # 2. Stats calculation
        start_date = None
        days = self.request.GET.get('days')
        if days and days.isdigit():
            start_date = date.today() - timedelta(days=int(days))
        
        stats = get_customer_debt_stats(customer_id, start_date)
        
        from orders.models import OrderStatus, PurchaseOrder, SalesOrder
        context.update({
            'customer': customer,
            'grouped_entries': grouped_entries,
            'receivable_balance': stats['receivable_balance'],
            'payable_balance': stats['payable_balance'],
            'net_balance': stats['net_balance'],
            'draft_sales': SalesOrder.objects.filter(customer=customer, status=OrderStatus.DRAFT).count(),
            'draft_purchases': PurchaseOrder.objects.filter(supplier=customer, status=OrderStatus.DRAFT).count(),
            'current_days': days or ''
        })
        return context


class ExportDebtHistoryView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Excel export with full audit trail."""
    def test_func(self):
        return self.request.user.role == 'ADMIN'

    def get(self, request, customer_id):
        customer = get_object_or_404(Customer, pk=customer_id)
        qs = DebtEntry.objects.filter(customer=customer).select_related(
            'sales_order', 'purchase_order', 'settlement', 'parent_entry'
        )
        
        days = request.GET.get('days')
        time_label = "Tất cả"
        if days and days.isdigit():
            start_date = date.today() - timedelta(days=int(days))
            qs = qs.filter(entry_date__gte=start_date)
            time_label = f"Trong {days} ngày"

        qs = qs.order_by('-entry_date', '-created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử công nợ"

        headers = ["Ngày ghi", "Nội dung / Chứng từ", "Mã đơn", "Giá trị đơn mua", "Giá trị đơn bán", "Đã thanh toán", "Khoản phải thu (+)", "Khoản phải trả (-)"]
        ws.append(headers)
        
        # Styling
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for entry in qs:
            po_val = 0; so_val = 0; paid = 0; r_bal = 0; p_bal = 0
            
            # Order Code logic
            code = entry.sales_order.code if entry.sales_order else (entry.purchase_order.code if entry.purchase_order else "")
            if not code and entry.parent_entry:
                parent = entry.parent_entry
                code = f"Trả cho {parent.sales_order.code if parent.sales_order else parent.purchase_order.code}"

            if entry.is_settlement:
                paid = float(entry.amount)
                type_lbl = "[Thu]" if entry.account_type == AccountType.RECEIVABLE else "[Chi]"
                if entry.settlement:
                    s_date = entry.settlement.payment_date.strftime("%d/%m/%Y")
                    raw_note = entry.note.replace("Quyết toán (FIFO): ", "") if entry.note else ""
                    note = f"{type_lbl} Quyết toán #{entry.settlement.id} ngày {s_date} - {raw_note}"
                else:
                    note = f"{type_lbl} {entry.note or ''}"
                
                if entry.account_type == AccountType.RECEIVABLE: r_bal = -paid
                else: p_bal = -paid
            else:
                note = entry.note or ""
                if entry.account_type == AccountType.RECEIVABLE:
                    so_val = float(entry.amount); r_bal = so_val
                else:
                    po_val = float(entry.amount); p_bal = po_val

            ws.append([
                entry.entry_date.strftime("%d/%m/%Y %H:%M") if entry.entry_date else entry.created_at.strftime("%d/%m/%Y %H:%M"),
                note, code, po_val, so_val, paid, r_bal, p_bal
            ])

        # Auto-adjust columns
        for col in ws.columns:
            column = col[0].column_letter
            ws.column_dimensions[column].width = max(len(str(cell.value) or "") for cell in col) + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="CongNo_{customer.name}_{date.today()}.xlsx"'
        wb.save(response)
        return response


class EntryPaymentView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Direct payment for a single debt entry."""
    def test_func(self):
        return self.request.user.role == 'ADMIN'
    
    def get(self, request, pk):
        entry = get_object_or_404(DebtEntry, pk=pk)
        form = EntryPaymentForm(initial={'amount': entry.remaining_amount, 'payment_date': date.today()})
        return render(request, 'debt/entry_payment_form.html', {'form': form, 'entry': entry})

    def post(self, request, pk):
        entry = get_object_or_404(DebtEntry, pk=pk)
        form = EntryPaymentForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                s = Settlement.objects.create(
                    customer=entry.customer, account_type=entry.account_type,
                    amount_paid=form.cleaned_data['amount'], 
                    payment_date=form.cleaned_data['payment_date'],
                    note=form.cleaned_data['note']
                )
                DebtEntry.objects.create(
                    customer=entry.customer, account_type=entry.account_type,
                    parent_entry=entry, settlement=s, amount=s.amount_paid, is_settlement=True,
                    note=f"Thanh toán đơn {entry.sales_order.code if entry.sales_order else entry.purchase_order.code}",
                    entry_date=s.payment_date
                )
            messages.success(request, f"Đã thanh toán {s.amount_paid} cho {entry.customer.name}")
            return redirect('debt:customer-debt', customer_id=entry.customer.pk)
        return render(request, 'debt/entry_payment_form.html', {'form': form, 'entry': entry})


class SettlementCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Bulk FIFO settlement creation."""
    model = Settlement
    form_class = SettlementForm
    template_name = 'debt/settlement_form.html'
    
    def test_func(self):
        return self.request.user.role == 'ADMIN'
    
    def get_initial(self):
        return {
            'payment_date': date.today(),
            'customer': self.request.GET.get('customer'),
            'account_type': self.request.GET.get('account_type')
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cid = self.request.GET.get('customer')
        if cid:
            stats = get_customer_debt_stats(cid)
            context.update({
                'current_customer': get_object_or_404(Customer, pk=cid),
                'net_balance': stats['net_balance'],
                'abs_net_balance': abs(stats['net_balance'])
            })
        return context

    def form_valid(self, form):
        with transaction.atomic():
            s = form.save()
            allocate_settlement_fifo(s, s.amount_paid)
            messages.success(self.request, f"Đã quyết toán {s.amount_paid} cho {s.customer.name}")
            return redirect('debt:settlement-success')


class SettlementSuccessView(LoginRequiredMixin, TemplateView):
    template_name = "debt/settlement_success.html"


class SettlementDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Safely delete a settlement and its linked allocations."""
    model = Settlement
    def test_func(self):
        return self.request.user.role == 'ADMIN'
        
    def get_success_url(self):
        messages.success(self.request, "Đã xóa phiếu quyết toán. Công nợ đã được khôi phục.")
        return reverse_lazy('debt:customer-debt', kwargs={'customer_id': self.object.customer.id})

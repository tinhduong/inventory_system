from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.views import View
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Warehouse, Product, StockItem
from .forms import WarehouseForm, ProductForm, ExcelImportForm

class WarehouseListView(LoginRequiredMixin, ListView):
    model = Warehouse
    template_name = 'catalog/warehouse_list.html'
    context_object_name = 'warehouses'

class WarehouseCreateView(LoginRequiredMixin, CreateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'catalog/warehouse_form.html'
    success_url = reverse_lazy('catalog:warehouse-list')

class WarehouseUpdateView(LoginRequiredMixin, UpdateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'catalog/warehouse_form.html'
    success_url = reverse_lazy('catalog:warehouse-list')

class WarehouseDetailView(LoginRequiredMixin, DetailView):
    model = Warehouse
    template_name = 'catalog/warehouse_detail.html'
    context_object_name = 'warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Lấy tồn kho tại kho này
        context['stocks'] = StockItem.objects.filter(warehouse=self.object, quantity__gt=0).select_related('product')
        
        # Lấy lịch sử nhập hàng gần đây liên quan đến kho này (đã xác nhận)
        from orders.models import PurchaseOrder
        context['recent_purchases'] = PurchaseOrder.objects.filter(
            warehouse=self.object, 
            status='CONFIRMED'
        ).order_by('-updated_at')[:10]
        
        return context

class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'catalog/product_list.html'
    context_object_name = 'products'

class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'catalog/product_form.html'
    success_url = reverse_lazy('catalog:product-list')

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'catalog/product_form.html'
    success_url = reverse_lazy('catalog:product-list')

class StockListView(LoginRequiredMixin, ListView):
    model = StockItem
    template_name = 'catalog/stock_list.html'
    context_object_name = 'stocks'
    
    def get_queryset(self):
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        
        self.warehouse_id = self.request.GET.get('warehouse', 'all')
        
        if self.warehouse_id and self.warehouse_id != 'all':
            # Nếu có chọn kho cụ thể
            return StockItem.objects.filter(
                warehouse_id=self.warehouse_id
            ).select_related('product', 'warehouse').order_by('product__name')
        else:
            # "Tất cả kho" hoặc chưa chọn (mặc định cho xem tất cả)
            from django.db.models import Q
            return Product.objects.annotate(
                total_quantity=Coalesce(Sum('stocks__quantity'), 0),
                total_held=Coalesce(Sum('stocks__held_quantity'), 0),
                total_incoming=Coalesce(Sum('stocks__incoming_quantity'), 0)
            ).filter(
                Q(total_quantity__gt=0) | Q(total_held__gt=0) | Q(total_incoming__gt=0)
            ).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouses'] = Warehouse.objects.all()
        # Ensure current_warehouse is in context
        context['current_warehouse'] = getattr(self, 'warehouse_id', 'all')
        return context

class StockHeldDetailView(LoginRequiredMixin, ListView):
    template_name = 'catalog/stock_held_detail.html'
    context_object_name = 'lines'

    def get_queryset(self):
        from orders.models import SalesOrderLine, OrderStatus
        product_id = self.request.GET.get('product')
        warehouse_id = self.request.GET.get('warehouse')
        
        self.product = Product.objects.get(id=product_id)
        
        qs = SalesOrderLine.objects.filter(
            order__status=OrderStatus.DRAFT,
            product_id=product_id
        ).select_related('order', 'order__customer', 'order__employee', 'order__warehouse')
        
        if warehouse_id and warehouse_id != 'all':
            qs = qs.filter(order__warehouse_id=warehouse_id)
            self.warehouse = Warehouse.objects.get(id=warehouse_id)
        else:
            self.warehouse = None
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        context['warehouse'] = self.warehouse
        return context

class ExportProductsView(LoginRequiredMixin, View):
    def get(self, request):
        products = Product.objects.all().order_by('code')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh muc san pham"
        
        headers = ["Ma SKU", "Ten san pham", "Mo ta", "Don vi tinh", "Nguong canh bao"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        for product in products:
            ws.append([
                product.code,
                product.name,
                product.description or "",
                product.unit,
                product.min_stock
            ])
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[column].width = max_length + 2
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="DanhSachSanPham.xlsx"'
        wb.save(response)
        return response

class ImportProductsView(LoginRequiredMixin, View):
    def get(self, request):
        form = ExcelImportForm()
        return render(request, 'catalog/product_import.html', {'form': form})
        
    def post(self, request):
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                
                # Bắt đầu từ row 2 để bỏ qua header
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not row[0]: continue # Bỏ qua dòng trống
                    
                    code = str(row[0]).strip()
                    name = str(row[1]).strip() if row[1] else ""
                    description = str(row[2]).strip() if row[2] else ""
                    unit = str(row[3]).strip() if row[3] else "Cái"
                    min_stock = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                    
                    if not code or not name:
                        error_count += 1
                        continue
                        
                    # Cập nhật nếu đã tồn tại, ngược lại tạo mới
                    obj, created = Product.objects.update_or_create(
                        code=code,
                        defaults={
                            'name': name,
                            'description': description,
                            'unit': unit,
                            'min_stock': min_stock
                        }
                    )
                    success_count += 1
                    
                messages.success(request, f"Import thành công: {success_count} sản phẩm. Lỗi: {error_count}")
                return redirect('catalog:product-list')
            except Exception as e:
                messages.error(request, f"Lỗi xử lý file Excel: {str(e)}")
        
        return render(request, 'catalog/product_import.html', {'form': form})

class CheckStockView(LoginRequiredMixin, View):
    def get(self, request):
        product_id = request.GET.get('product_id')
        warehouse_id = request.GET.get('warehouse_id')
        if not product_id or not warehouse_id:
            return JsonResponse({'available': 0})
        
        try:
            stock = StockItem.objects.get(product_id=product_id, warehouse_id=warehouse_id)
            available = stock.available_quantity
        except StockItem.DoesNotExist:
            available = 0
            
        return JsonResponse({'available': available})
